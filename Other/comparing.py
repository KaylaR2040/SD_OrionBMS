import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

MODULE_FILE = "decoded_can.xlsx"      # real thermistor module log
MYCAN_FILE  = "decoded_mycan.xlsx"    # your implementation log
OUT_FILE    = "can_validation_report.xlsx"

GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="F4CCCC")
BLUE = PatternFill("solid", fgColor="1F4E79")
GRAY = PatternFill("solid", fgColor="EDEDED")

WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="B7B7B7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


EXPECTED = {
    "18EEFF80": {
        "description": "J1939 Address Claim Broadcast",
        "length": 8,
        "fields": [
            "Unique identifier for J1939 compatible address claim",
            "BMS Target Address",
            "Thermistor Module Number",
            "0x40 (Constant)",
            "0x1E (Constant)",
            "0x90 (Constant)",
        ],
        "checksum": False,
    },
    "1839F380": {
        "description": "Thermistor Module -> BMS Broadcast",
        "length": 8,
        "fields": [
            "Thermistor Module Number",
            "Lowest thermistor value",
            "Highest thermistor value",
            "Average thermistor value",
            "Thermistors enabled count",
            "Fault bit in Byte 5",
            "Highest Thermistor ID",
            "Lowest Thermistor ID",
            "Checksum Byte",
        ],
        "checksum": True,
    },
    "1838F380": {
        "description": "Thermistor General Broadcast",
        "length": 8,
        "fields": [
            "Thermistor ID relative to all configured modules",
            "Current thermistor value",
            "Thermistor ID relative to this module",
            "Lowest thermistor value",
            "Highest thermistor value",
            "Highest Thermistor ID",
            "Lowest Thermistor ID",
        ],
        "checksum": False,
    },
    "80": {
        "description": "Legacy Broadcast Message - RESERVED",
        "length": 4,
        "fields": ["RESERVED FOR LEGACY EQUIPMENT"],
        "checksum": False,
    }
}


def normalize(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    s = s.replace("°C", "").replace("(signed)", "").replace(",", " ")
    s = s.replace("  ", " ")
    return s.lower()


def hex_to_int(v):
    if pd.isna(v):
        return None
    s = str(v).strip()
    if s.startswith("0x") or s.startswith("0X"):
        return int(s, 16)
    return int(s)


def read_decoded(path):
    return pd.read_excel(path, sheet_name="Decoded")


def read_raw(path):
    return pd.read_excel(path, sheet_name="Raw_Bytes")


def checksum_1839(row):
    vals = []
    for i in range(1, 8):
        vals.append(hex_to_int(row[f"Byte {i}"]))
    expected = (sum(vals) + 0x39 + 0x08) & 0xFF
    actual = hex_to_int(row["Byte 8"])
    return expected, actual, expected == actual


def extract_presence(df):
    counts = df["CAN ID"].astype(str).value_counts().to_dict()
    return counts


def row_status(ok):
    return GREEN if ok else RED


def soft_status(ok, warn=False):
    if ok:
        return GREEN
    if warn:
        return YELLOW
    return RED


module_dec = read_decoded(MODULE_FILE)
mycan_dec = read_decoded(MYCAN_FILE)
module_raw = read_raw(MODULE_FILE)
mycan_raw = read_raw(MYCAN_FILE)

module_counts = extract_presence(module_dec)
mycan_counts = extract_presence(mycan_dec)

# ---------------------------------------------------------
# Build summary results
# ---------------------------------------------------------
summary_rows = []
for can_id, spec in EXPECTED.items():
    m_count = module_counts.get(can_id, 0)
    y_count = mycan_counts.get(can_id, 0)

    present_module = m_count > 0
    present_mycan = y_count > 0

    if present_module and present_mycan:
        status = "MATCHING PRESENCE"
    elif present_module and not present_mycan:
        status = "MISSING FROM MYCAN"
    elif not present_module and present_mycan:
        status = "ONLY IN MYCAN"
    else:
        status = "MISSING FROM BOTH"

    summary_rows.append({
        "CAN ID": can_id,
        "Expected Description": spec["description"],
        "Module Count": m_count,
        "MyCAN Count": y_count,
        "Status": status
    })

summary_df = pd.DataFrame(summary_rows)

# ---------------------------------------------------------
# Detailed 1839 check
# ---------------------------------------------------------
detail_rows = []

for source_name, df in [("MODULE", module_dec), ("MYCAN", mycan_dec)]:
    df_1839 = df[df["CAN ID"].astype(str) == "1839F380"].copy()

    for idx, row in df_1839.head(25).iterrows():
        exp_crc, act_crc, crc_ok = checksum_1839(row)

        detail_rows.append({
            "Source": source_name,
            "Line #": row["Line #"],
            "CAN ID": row["CAN ID"],
            "Length": row["Length"],
            "Byte 1": row["Byte 1"],
            "Byte 2": row["Byte 2"],
            "Byte 3": row["Byte 3"],
            "Byte 4": row["Byte 4"],
            "Byte 5": row["Byte 5"],
            "Byte 6": row["Byte 6"],
            "Byte 7": row["Byte 7"],
            "Byte 8": row["Byte 8"],
            "Expected CRC": f"0x{exp_crc:02X}",
            "Actual CRC": f"0x{act_crc:02X}" if act_crc is not None else "",
            "CRC OK": crc_ok
        })

detail_df = pd.DataFrame(detail_rows)

# ---------------------------------------------------------
# Field layout comparison by CAN ID
# ---------------------------------------------------------
def get_fields_seen(df, can_id):
    subset = df[df["CAN ID"].astype(str) == can_id]
    fields = set()

    for _, row in subset.iterrows():
        for col in df.columns:
            if str(col).startswith("Field"):
                val = normalize(row[col])
                if val:
                    fields.add(val)

        # include explicit columns too
        for special in ["Highest Thermistor ID", "Lowest Thermistor ID", "Checksum Byte"]:
            if special in df.columns and not pd.isna(row.get(special, None)):
                fields.add(normalize(special))

    return fields

layout_rows = []
for can_id, spec in EXPECTED.items():
    expected_fields = {normalize(x) for x in spec["fields"]}
    module_fields = get_fields_seen(module_dec, can_id)
    mycan_fields = get_fields_seen(mycan_dec, can_id)

    missing_in_mycan = sorted(expected_fields - mycan_fields)
    missing_in_module = sorted(expected_fields - module_fields)

    layout_rows.append({
        "CAN ID": can_id,
        "Expected Fields": ", ".join(sorted(expected_fields)),
        "Missing in Module": ", ".join(missing_in_module),
        "Missing in MyCAN": ", ".join(missing_in_mycan),
        "Module Field Coverage OK": len(missing_in_module) == 0,
        "MyCAN Field Coverage OK": len(missing_in_mycan) == 0,
    })

layout_df = pd.DataFrame(layout_rows)

# ---------------------------------------------------------
# Workbook output
# ---------------------------------------------------------
wb = Workbook()
ws1 = wb.active
ws1.title = "Summary"

def write_table(ws, start_row, df, title):
    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).fill = BLUE
    ws.cell(start_row, 1).font = WHITE_FONT
    ws.cell(start_row, 1).alignment = CENTER

    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(start_row + 1, c, col)
        cell.fill = GRAY
        cell.font = BOLD
        cell.alignment = CENTER
        cell.border = BORDER

    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 2):
        for c_idx, col in enumerate(df.columns, start=1):
            cell = ws.cell(r_idx, c_idx, row[col])
            cell.border = BORDER
            cell.alignment = LEFT

            if "Status" in df.columns and col == "Status":
                text = str(row[col])
                if "MATCH" in text or "OK" in text:
                    cell.fill = GREEN
                elif "MISSING" in text or "ONLY" in text:
                    cell.fill = YELLOW
                else:
                    cell.fill = RED

            if col == "CRC OK":
                cell.fill = GREEN if row[col] else RED
                cell.alignment = CENTER

            if col in ["Module Field Coverage OK", "MyCAN Field Coverage OK"]:
                cell.fill = GREEN if row[col] else RED
                cell.alignment = CENTER

# Summary sheet
write_table(ws1, 1, summary_df, "CAN ID Presence Comparison")
write_table(ws1, len(summary_df) + 5, layout_df, "Expected Field Coverage vs Each File")

# 1839 CRC sheet
ws2 = wb.create_sheet("1839_CRC_Check")
write_table(ws2, 1, detail_df, "1839F380 CRC Validation")

# Notes sheet
ws3 = wb.create_sheet("Interpretation")
notes = [
    ["A / MODULE means the real thermistor module log."],
    ["B / MYCAN means the log from your implementation."],
    ["MISSING FROM MYCAN means the real module sends that CAN frame but your code did not."],
    ["For 1839F380, CRC OK means Byte 8 matched: (Byte1+...+Byte7+0x39+0x08) & 0xFF."],
    ["If 1839F380 matches but 18EEFF80 and 1838F380 are missing, then your implementation is only partially matching the module."],
    ["Header-column differences in Excel do not necessarily mean the CAN frame itself is wrong."],
]
for i, row in enumerate(notes, start=1):
    c = ws3.cell(i, 1, row[0])
    c.border = BORDER
    c.alignment = LEFT

# autosize
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

wb.save(OUT_FILE)

print(f"Created {OUT_FILE}")