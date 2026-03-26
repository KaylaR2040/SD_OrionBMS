import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE = "22.txt"
OUTPUT_FILE = "22.xlsx"

# Only decode these CAN IDs
TARGET_IDS = {
    "18EEFF80",   # J1939 Address Claim Broadcast
    "1839F380",   # Thermistor Module -> BMS Broadcast
    "1838F380",   # Thermistor General Broadcast
    "00000080",   # Legacy Broadcast Message - RESERVED
    "80",         # in case it appears short
}

# Reference info based on your screenshot
REFERENCE_INFO = {
    "18EEFF80": {
        "Description": "J1939 Address Claim Broadcast",
        "Freq [ms]": 200,
        "Length": 8,
    },
    "1839F380": {
        "Description": "Thermistor Module -> BMS Broadcast",
        "Freq [ms]": 100,
        "Length": 8,
    },
    "1838F380": {
        "Description": "Thermistor General Broadcast",
        "Freq [ms]": 100,
        "Length": 8,
    },
    "00000080": {
        "Description": "Legacy Broadcast Message - RESERVED",
        "Freq [ms]": 100,
        "Length": 4,
    },
    "80": {
        "Description": "Legacy Broadcast Message - RESERVED",
        "Freq [ms]": 100,
        "Length": 4,
    },
}


# ============================================================
# HELPERS
# ============================================================

def to_signed_8bit(value: int) -> int:
    """Convert unsigned byte 0..255 to signed int8."""
    return value - 256 if value > 127 else value


def parse_can_line(line: str):
    """
    Parse slcan/lawicel style lines like:
      t6B380057000664179B31547B
      x1838F38180053150315150303

    Assumptions:
    - t = standard frame, 3 hex CAN ID, 1 hex DLC
    - x = extended frame, 8 hex CAN ID, 1 hex DLC
    - remaining data = 2 * DLC hex chars
    - anything left after that is ignored (often timestamp)
    """
    line = line.strip()
    if not line:
        return None

    prefix = line[0].lower()
    if prefix not in ("t", "x"):
        return None

    if prefix == "t":
        can_id_len = 3
    else:
        can_id_len = 8

    if len(line) < 1 + can_id_len + 1:
        return None

    can_id = line[1:1 + can_id_len].upper()
    dlc_char = line[1 + can_id_len]
    if dlc_char not in "0123456789ABCDEFabcdef":
        return None

    dlc = int(dlc_char, 16)
    data_start = 1 + can_id_len + 1
    data_end = data_start + (2 * dlc)
    if len(line) < data_end:
        return None

    data_hex = line[data_start:data_end]
    try:
        data_bytes = [int(data_hex[i:i+2], 16) for i in range(0, len(data_hex), 2)]
    except ValueError:
        return None

    return {
        "raw_line": line,
        "prefix": prefix,
        "can_id": can_id,
        "dlc": dlc,
        "data_bytes": data_bytes
    }


def normalize_can_id(can_id: str) -> str:
    """
    Make standard IDs and extended IDs easier to compare.
    """
    can_id = can_id.upper()
    if len(can_id) == 3 and can_id == "80":
        return "80"
    return can_id


def decode_frame(parsed: dict, line_number: int):
    can_id = normalize_can_id(parsed["can_id"])
    data = parsed["data_bytes"]
    info = REFERENCE_INFO.get(can_id, {})

    # Pad to 8 bytes so indexing is safe
    padded = data + [None] * (8 - len(data))
    b1, b2, b3, b4, b5, b6, b7, b8 = padded[:8]

    # Generic raw byte columns
    row = {
        "Line #": line_number,
        "Raw Line": parsed["raw_line"],
        "CAN ID": can_id,
        "Description": info.get("Description", ""),
        "Freq [ms]": info.get("Freq [ms]", ""),
        "Length": parsed["dlc"],
        "Byte 1": f"0x{b1:02X}" if b1 is not None else "",
        "Byte 2": f"0x{b2:02X}" if b2 is not None else "",
        "Byte 3": f"0x{b3:02X}" if b3 is not None else "",
        "Byte 4": f"0x{b4:02X}" if b4 is not None else "",
        "Byte 5": f"0x{b5:02X}" if b5 is not None else "",
        "Byte 6": f"0x{b6:02X}" if b6 is not None else "",
        "Byte 7": f"0x{b7:02X}" if b7 is not None else "",
        "Byte 8": f"0x{b8:02X}" if b8 is not None else "",
    }

    # --------------------------------------------------------
    # Decode exactly based on the table you showed
    # --------------------------------------------------------

    if can_id == "18EEFF80":
        row.update({
            "Field 1": "Unique identifier for J1939 compatible address claim",
            "Value 1": f"0x{b1:02X}" if b1 is not None else "",
            "Field 2": "BMS Target Address",
            "Value 2": f"0x{b4:02X}" if b4 is not None else "",
            "Field 3": "Thermistor Module Number",
            "Value 3": f"0x{b5:02X}" if b5 is not None else "",
            "Field 4": "Constant Byte 6",
            "Value 4": f"0x{b6:02X}" if b6 is not None else "",
            "Field 5": "Constant Byte 7",
            "Value 5": f"0x{b7:02X}" if b7 is not None else "",
            "Field 6": "Constant Byte 8",
            "Value 6": f"0x{b8:02X}" if b8 is not None else "",
        })

    elif can_id == "1839F380":
        # Note #3: bit 7 / 0x80 indicates fault present in byte 5
        therm_count = b5 & 0x7F if b5 is not None else None
        fault_present = 1 if (b5 is not None and (b5 & 0x80)) else 0

        row.update({
            "Field 1": "Thermistor Module Number",
            "Value 1": b1 if b1 is not None else "",
            "Field 2": "Lowest thermistor value (°C, signed)",
            "Value 2": to_signed_8bit(b2) if b2 is not None else "",
            "Field 3": "Highest thermistor value (°C, signed)",
            "Value 3": to_signed_8bit(b3) if b3 is not None else "",
            "Field 4": "Average thermistor value (°C, signed)",
            "Value 4": to_signed_8bit(b4) if b4 is not None else "",
            "Field 5": "Thermistors enabled count",
            "Value 5": therm_count if therm_count is not None else "",
            "Field 6": "Fault bit in Byte 5",
            "Value 6": fault_present,
            "Highest Thermistor ID": b6 if b6 is not None else "",
            "Lowest Thermistor ID": b7 if b7 is not None else "",
            "Checksum Byte": f"0x{b8:02X}" if b8 is not None else "",
        })

    elif can_id == "1838F380":
        # Notes from screenshot:
        # byte1 = Thermistor ID relative to all configured Thermistor Modules
        # byte2 = Thermistor value (signed, °C)
        # byte3 = Thermistor ID relative to this thermistor module
        # byte4 = Lowest thermistor value (signed, °C)
        # byte5 = Highest thermistor value (signed, °C)
        # byte6 = Highest thermistor ID on the module
        # byte7 = Lowest thermistor ID on the module
        row.update({
            "Field 1": "Thermistor ID relative to all configured modules",
            "Value 1": b1 if b1 is not None else "",
            "Field 2": "Current thermistor value (°C, signed)",
            "Value 2": to_signed_8bit(b2) if b2 is not None else "",
            "Field 3": "Thermistor ID relative to this module",
            "Value 3": b3 if b3 is not None else "",
            "Field 4": "Lowest thermistor value (°C, signed)",
            "Value 4": to_signed_8bit(b4) if b4 is not None else "",
            "Field 5": "Highest thermistor value (°C, signed)",
            "Value 5": to_signed_8bit(b5) if b5 is not None else "",
            "Highest Thermistor ID": b6 if b6 is not None else "",
            "Lowest Thermistor ID": b7 if b7 is not None else "",
            "Byte 8 / Extra": f"0x{b8:02X}" if b8 is not None else "",
        })

    elif can_id in ("80", "00000080"):
        row.update({
            "Field 1": "Reserved legacy message",
            "Value 1": "",
        })

    return row


# ============================================================
# MAIN
# ============================================================

def main():
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        raise FileNotFoundError(f"Could not find input file: {INPUT_FILE}")

    decoded_rows = []
    raw_rows = []

    with input_path.open("r", encoding="utf-8") as f:
        for line_number, line in enumerate(f, start=1):
            parsed = parse_can_line(line)
            if not parsed:
                continue

            can_id = normalize_can_id(parsed["can_id"])
            if can_id not in TARGET_IDS:
                continue

            decoded_rows.append(decode_frame(parsed, line_number))

            # raw sheet row
            data = parsed["data_bytes"] + [None] * (8 - len(parsed["data_bytes"]))
            raw_rows.append({
                "Line #": line_number,
                "Raw Line": parsed["raw_line"],
                "CAN ID": can_id,
                "DLC": parsed["dlc"],
                "Byte 1": f"0x{data[0]:02X}" if data[0] is not None else "",
                "Byte 2": f"0x{data[1]:02X}" if data[1] is not None else "",
                "Byte 3": f"0x{data[2]:02X}" if data[2] is not None else "",
                "Byte 4": f"0x{data[3]:02X}" if data[3] is not None else "",
                "Byte 5": f"0x{data[4]:02X}" if data[4] is not None else "",
                "Byte 6": f"0x{data[5]:02X}" if data[5] is not None else "",
                "Byte 7": f"0x{data[6]:02X}" if data[6] is not None else "",
                "Byte 8": f"0x{data[7]:02X}" if data[7] is not None else "",
            })

    if not decoded_rows:
        print("No matching CAN IDs were found.")
        return

    decoded_df = pd.DataFrame(decoded_rows)
    raw_df = pd.DataFrame(raw_rows)

    # Reference sheet to mirror your screenshot at a high level
    reference_rows = [
        {
            "CAN ID": "18EEFF80",
            "Description": "J1939 Address Claim Broadcast",
            "Freq [ms]": 200,
            "Length": 8,
            "Byte #1": "Unique identifier for J1939 compatible address claim",
            "Byte #4": "BMS Target Address",
            "Byte #5": "Thermistor Module Number",
            "Byte #6": "0x40 (Constant)",
            "Byte #7": "0x1E (Constant)",
            "Byte #8": "0x90 (Constant)",
        },
        {
            "CAN ID": "1839F380",
            "Description": "Thermistor Module -> BMS Broadcast",
            "Freq [ms]": 100,
            "Length": 8,
            "Byte #1": "Thermistor Module Number",
            "Byte #2": "Lowest thermistor value (signed)",
            "Byte #3": "Highest thermistor value (signed)",
            "Byte #4": "Average thermistor value (signed)",
            "Byte #5": "Enabled count + fault bit",
            "Byte #6": "Highest thermistor ID",
            "Byte #7": "Lowest thermistor ID",
            "Byte #8": "Checksum",
        },
        {
            "CAN ID": "1838F380",
            "Description": "Thermistor General Broadcast",
            "Freq [ms]": 100,
            "Length": 8,
            "Byte #1": "Thermistor ID relative to all configured modules",
            "Byte #2": "Current thermistor value (signed)",
            "Byte #3": "Thermistor ID relative to this module",
            "Byte #4": "Lowest thermistor value (signed)",
            "Byte #5": "Highest thermistor value (signed)",
            "Byte #6": "Highest thermistor ID",
            "Byte #7": "Lowest thermistor ID",
            "Byte #8": "Extra / raw",
        },
        {
            "CAN ID": "80",
            "Description": "Legacy Broadcast Message - RESERVED",
            "Freq [ms]": 100,
            "Length": 4,
            "Byte #1": "Reserved",
        }
    ]
    reference_df = pd.DataFrame(reference_rows)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        decoded_df.to_excel(writer, index=False, sheet_name="Decoded")
        raw_df.to_excel(writer, index=False, sheet_name="Raw_Bytes")
        reference_df.to_excel(writer, index=False, sheet_name="Reference")

    # ========================================================
    # Pretty formatting
    # ========================================================
    wb = load_workbook(OUTPUT_FILE)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap

        # Auto width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    wb.save(OUTPUT_FILE)
    print(f"Done. Wrote {OUTPUT_FILE}")


if __name__ == "__main__":
    main()